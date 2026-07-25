"""Lecture et écriture de fichiers Excel/CSV pour l'import et l'export Odoo."""

from __future__ import annotations

import csv
from pathlib import Path

from odoo_mcp.odoo_client import OdooError


def read_rows(path: Path, sheet: str = "", header_row: int = 1) -> tuple[list[str], list[list]]:
    """Retourne (en-têtes, lignes de données) d'un .xlsx/.xlsm/.csv."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # openpyxl est dans les dépendances, mais au cas où
            raise OdooError("openpyxl manquant — réinstalle le serveur.") from exc
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet] if sheet else wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        header: list[str] = []
        for _ in range(header_row):
            row = next(it, None)
            if row is None:
                raise OdooError(f"{path.name} : ligne d'en-tête {header_row} introuvable.")
            header = [str(h).strip() if h is not None else "" for h in row]
        rows = [list(r) for r in it
                if any(v is not None and str(v).strip() != "" for v in r)]
        wb.close()
        return header, rows

    if path.suffix.lower() == ".csv":
        for enc in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                with path.open(encoding=enc, newline="") as fh:
                    sample = fh.read(8192)
                    fh.seek(0)
                    try:
                        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                    except csv.Error:
                        dialect = csv.excel
                    reader = csv.reader(fh, dialect)
                    header = []
                    for _ in range(header_row):
                        header = [h.strip() for h in next(reader)]
                    rows = [r for r in reader if any(str(v).strip() for v in r)]
                return header, rows
            except UnicodeDecodeError:
                continue
        raise OdooError(f"Encodage de {path.name} non reconnu.")

    raise OdooError(f"Format non géré : {path.suffix} (attendu .xlsx, .xlsm ou .csv)")


def inspect_summary(header: list[str], rows: list[list]) -> dict:
    """Structure du fichier : remplissage, exemples et valeurs distinctes par colonne."""
    cols = []
    for i, h in enumerate(header):
        if not h:
            continue
        vals = [r[i] for r in rows if i < len(r) and r[i] not in (None, "")]
        distinct = {str(v) for v in vals}
        col = {
            "colonne": h,
            "rempli": f"{len(vals)}/{len(rows)}",
            "exemples": [str(v)[:40] for v in vals[:3]],
        }
        if 0 < len(distinct) <= 8 and len(vals) > 8:
            col["valeurs_distinctes"] = sorted(v[:30] for v in distinct)
        cols.append(col)

    summary: dict = {"lignes": len(rows), "colonnes": cols}
    for key in ("id", "ID", "External ID", "Code"):
        if key in header:
            i = header.index(key)
            ids = [str(r[i]) for r in rows if i < len(r) and r[i] not in (None, "")]
            summary["identifiants"] = {
                "colonne": key, "uniques": len(set(ids)),
                "doublons": len(ids) - len(set(ids)),
            }
            break
    return summary


def build(header: list[str], rows: list[list],
          mapping: dict) -> tuple[list[str], list[list[str]]]:
    """Applique le mapping colonnes → champs et retourne (fields, data) pour load()."""
    cols = mapping.get("_columns") or {h: h for h in header if h}
    consts = mapping.get("_constants") or {}
    repl = mapping.get("_replace") or {}

    idx = {h: i for i, h in enumerate(header) if h}
    fields, sources = [], []
    for src, target in cols.items():
        if not target or src not in idx:
            continue
        fields.append(target)
        sources.append(idx[src])
    fields += list(consts)

    out = []
    for r in rows:
        line = []
        for field, i in zip(fields, sources):
            v = r[i] if i < len(r) else None
            v = "" if v is None else v
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            v = str(v).strip()
            table = repl.get(field)
            if table and v in table:
                v = table[v]
            line.append(v)
        line += [str(consts[f]) for f in consts]
        out.append(line)
    return fields, out


def flatten(value):
    """Aplati une valeur Odoo pour l'export : many2one → libellé, x2many → cardinalité."""
    if isinstance(value, (list, tuple)):
        if len(value) == 2 and isinstance(value[0], int):
            return value[1]
        return ", ".join(str(v) for v in value) if len(value) <= 6 else f"[{len(value)}]"
    if value is False:
        return ""
    return value


def write_table(path: Path, headers: list[str], rows: list[list]) -> None:
    """Écrit un .xlsx (mise en forme légère) ou un .csv selon l'extension."""
    if path.suffix.lower() == ".xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F3864")
        for row in rows:
            ws.append(row)
        for j, h in enumerate(headers, 1):
            width = max([len(str(h))] + [len(str(r[j - 1])) for r in rows[:200] if len(r) >= j])
            ws.column_dimensions[get_column_letter(j)].width = min(width + 3, 50)
        ws.freeze_panes = "A2"
        wb.save(path)
    elif path.suffix.lower() == ".csv":
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            w = csv.writer(fh, delimiter=";")
            w.writerow(headers)
            w.writerows(rows)
    else:
        raise OdooError(f"Extension non gérée : {path.suffix} (attendu .xlsx ou .csv)")
